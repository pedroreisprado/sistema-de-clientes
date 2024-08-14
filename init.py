import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from PIL import ImageTk, Image
from datetime import datetime
import os

# Cores do tema moderno estilo Discord
BACKGROUND_COLOR = "#2c2f33"
FOREGROUND_COLOR = "#ffffff"
HIGHLIGHT_COLOR = "#7289da"
ENTRY_BG_COLOR = "#23272a"
ENTRY_FG_COLOR = "#ffffff"
BUTTON_BG_COLOR = "#7289da"
BUTTON_FG_COLOR = "#ffffff"
TREEVIEW_BG_COLOR = "#ffffff"  # Cor de fundo do Treeview
TREEVIEW_FG_COLOR = "#000000"  # Cor do texto do Treeview
HEADER_BG_COLOR = "#ffffff"  # Cor de fundo do cabeçalho
HEADER_FG_COLOR = "#000000"  # Cor do texto do cabeçalho

# Verifica se o arquivo base.xlsx existe; caso contrário, cria um novo
if not os.path.exists("aux_files/base.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clientes"
    sheet["A1"] = "Data"
    sheet["B1"] = "Nome do Cliente"
    sheet["C1"] = "Telefone"
    workbook.save("aux_files/base.xlsx")

def save_to_excel(name, phone):
    try:
        # Carrega o arquivo Excel
        workbook = load_workbook("aux_files/base.xlsx")
        sheet = workbook.active

        # Adiciona os dados na próxima linha vazia
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.cell(row=row, column=2).value = name
        sheet.cell(row=row, column=3).value = phone

        # Salva o arquivo
        workbook.save("aux_files/base.xlsx")
        return True
    except Exception as e:
        print("Erro ao salvar no Excel:", e)
        return False

def update_treeview():
    # Limpa a Treeview antes de atualizar
    for row in treeview.get_children():
        treeview.delete(row)

    # Carrega os dados do Excel
    workbook = load_workbook("aux_files/base.xlsx")
    sheet = workbook.active

    # Adiciona os últimos 5 registros à Treeview, pulando a primeira linha (cabeçalho)
    for row in range(sheet.max_row, max(sheet.max_row - 10, 1), -1):
        date = sheet.cell(row=row, column=1).value
        name = sheet.cell(row=row, column=2).value
        phone = sheet.cell(row=row, column=3).value

        treeview.insert("", "end", values=(date, name, phone))

def submit(event=None):
    name = entry_name.get()
    phone = entry_phone.get()

    if name and phone:
        if save_to_excel(name, phone):
            messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")
            entry_name.delete(0, tk.END)
            entry_phone.delete(0, tk.END)
            update_treeview()
        else:
            messagebox.showerror("Erro", "Falha ao salvar os dados.")
    else:
        messagebox.showwarning("Atenção", "Preencha todos os campos.")

def show_main_screen():
    welcome_screen.destroy()

    # Cria a tela principal
    global main_screen
    main_screen = tk.Tk()
    main_screen.title("Informações do Cliente")
    main_screen.geometry("820x620")
    main_screen.configure(bg=BACKGROUND_COLOR)

    # Labels e Entry para Nome e Telefone
    tk.Label(main_screen, text="Nome do Cliente:", fg=FOREGROUND_COLOR, bg=BACKGROUND_COLOR, font=("Arial", 14)).pack(pady=10)
    global entry_name
    entry_name = tk.Entry(main_screen, bg=ENTRY_BG_COLOR, fg=ENTRY_FG_COLOR, font=("Arial", 12))
    entry_name.pack(pady=5)

    tk.Label(main_screen, text="Telefone:", fg=FOREGROUND_COLOR, bg=BACKGROUND_COLOR, font=("Arial", 14)).pack(pady=10)
    global entry_phone
    entry_phone = tk.Entry(main_screen, bg=ENTRY_BG_COLOR, fg=ENTRY_FG_COLOR, font=("Arial", 12))
    entry_phone.pack(pady=5)

    # Botão de Enviar
    send_button = tk.Button(main_screen, text="Enviar", command=submit, bg=BUTTON_BG_COLOR, fg=BUTTON_FG_COLOR, font=("Arial", 12), relief="flat")
    send_button.pack(pady=20)

    # Vincula a tecla Enter para acionar o botão Enviar
    main_screen.bind('<Return>', submit)

    # Treeview para mostrar os últimos clientes adicionados
    tk.Label(main_screen, text="Últimos Clientes Adicionados:", fg=FOREGROUND_COLOR, bg=BACKGROUND_COLOR, font=("Arial", 14)).pack(pady=10)

    columns = ("Data", "Nome do Cliente", "Telefone")
    global treeview
    treeview = ttk.Treeview(main_screen, columns=columns, show="headings", height=10)
    treeview.pack(pady=5)

    # Definindo o estilo para o Treeview
    style = ttk.Style()
    style.configure("Treeview", background=TREEVIEW_BG_COLOR, foreground=TREEVIEW_FG_COLOR, fieldbackground=TREEVIEW_BG_COLOR, font=("Arial", 12))
    style.configure("Treeview.Heading", background=HEADER_BG_COLOR, foreground=HEADER_FG_COLOR, font=("Arial", 14, "bold"))

    # Configurando as colunas
    treeview.column("Data", anchor=tk.CENTER, width=200)
    treeview.column("Nome do Cliente", anchor=tk.CENTER, width=300)
    treeview.column("Telefone", anchor=tk.CENTER, width=200)

    # Configurando os cabeçalhos
    treeview.heading("Data", text="Data")
    treeview.heading("Nome do Cliente", text="Nome do Cliente")
    treeview.heading("Telefone", text="Telefone")

    # Atualiza a treeview com os dados existentes
    update_treeview()

    main_screen.mainloop()

# Tela de boas-vindas
welcome_screen = tk.Tk()
welcome_screen.title("Boas-vindas")
welcome_screen.geometry("820x620")
welcome_screen.configure(bg=BACKGROUND_COLOR)

# Carrega a imagem do logo
logo = Image.open("assets/logo.png")
logo = logo.resize((820, 620), Image.Resampling.LANCZOS)
logo = ImageTk.PhotoImage(logo)

# Label para exibir o logo
tk.Label(welcome_screen, image=logo, bg=BACKGROUND_COLOR).pack(pady=20)

# Label para exibir o texto "Aguarde...":
tk.Label(welcome_screen, text="Aguarde...", font=("Arial", 16), fg=FOREGROUND_COLOR, bg=BACKGROUND_COLOR).pack(pady=10)

# Agenda para mostrar a tela principal após 3 segundos
welcome_screen.after(3000, show_main_screen)

# Inicia o loop da tela de boas-vindas
welcome_screen.mainloop()
